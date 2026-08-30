"""Text extraction from PDFs and XLSX files with page/sheet provenance."""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class RawPage:
    """A unit of raw extracted text with provenance (pdf page or xlsx sheet)."""

    page: int  # 1-based page number or sheet index
    label: str  # "page" or sheet name
    text: str
    is_table: bool = False
    meta: dict = field(default_factory=dict)


def _text_outside_tables(page, tables) -> str:
    """Plain text of a page excluding words that sit inside detected table cells.

    pdfplumber's extract_text() reads table cells in visual order, which
    interleaves multi-line rows across columns and scrambles label/value pairs
    (e.g. '49 | Minimum amount | Rs. 300' becomes 'Minimum 49 - Rs. 300.').
    Since detected tables are already emitted as clean '|'-joined rows, the
    plain-text part only needs the content OUTSIDE the table cells.
    `tables` are pdfplumber Table objects (duck-typed: need .cells).
    """
    try:
        cell_boxes = [tuple(c) for t in tables for c in (getattr(t, "cells", None) or [])]
    except Exception:  # noqa: BLE001
        return page.extract_text() or ""
    if not cell_boxes:
        return page.extract_text() or ""

    words = page.extract_words()
    outside = []
    for w in words:
        cx = (w["x0"] + w["x1"]) / 2
        cy = (w["top"] + w["bottom"]) / 2
        if any(bx0 <= cx <= bx1 and btop <= cy <= bbottom for bx0, btop, bx1, bbottom in cell_boxes):
            continue
        outside.append(w)
    if not outside:
        return ""

    # reassemble reading order: group words into lines by vertical position
    lines: dict[int, list] = {}
    for w in outside:
        lines.setdefault(round(w["top"] / 3), []).append(w)
    return "\n".join(
        " ".join(word["text"] for word in sorted(ws, key=lambda w: w["x0"]))
        for _, ws in sorted(lines.items())
    )


def _scheme_summary_facts(rows: list[list[str]]) -> str | None:
    """Render a 'SCHEME SUMMARY DOCUMENT' table as natural-language facts.

    Pipe-joined table rows embed poorly against natural questions ("what is
    the minimum SIP for Flexi Cap"), so — like the TER collapse below — we
    additionally emit one 'label of <fund>: value.' sentence per row. Rows
    look like [<row number>, <label>, <value>]; multi-line values are joined.
    """
    def _row_text(r: list[str]) -> str:
        return " ".join((c or "") for c in r).upper()

    if not rows or not any("SCHEME SUMMARY DOCUMENT" in _row_text(r) for r in rows[:3]):
        return None

    fund = ""
    for r in rows:
        cells = [(c or "").replace("\n", " ").strip() for c in r]
        cells = [c for c in cells if c]
        if any(c.lower().startswith("fund name") for c in cells):
            idx = next(i for i, c in enumerate(cells) if c.lower().startswith("fund name"))
            if idx + 1 < len(cells):
                fund = cells[idx + 1]
            break
    if not fund:
        return None

    facts: list[str] = []
    for r in rows:
        cells = [(c or "").replace("\n", " ").strip() for c in r]
        cells = [c for c in cells if c]
        if len(cells) < 2:
            continue
        if any("SCHEME SUMMARY DOCUMENT" in c.upper() for c in cells):
            continue
        if len(cells) >= 3 and re.fullmatch(r"\d{1,3}", cells[0]):
            label, value = cells[1], " ".join(cells[2:])
        else:
            label, value = cells[0], " ".join(cells[1:])
        if not label or not value:
            continue
        facts.append(f"{label} of {fund}: {value}")
    if not facts:
        return None
    header = f"Scheme Summary Document - {fund}"
    return "\n".join([header] + facts)


def extract_pdf(path: Path) -> list[RawPage]:
    pages: list[RawPage] = []
    try:
        import pdfplumber

        with pdfplumber.open(str(path)) as pdf:
            for i, p in enumerate(pdf.pages, start=1):
                try:
                    found = p.find_tables()
                except Exception:  # noqa: BLE001
                    found = []
                tables = []
                for t in found:
                    try:
                        extracted = t.extract()
                    except Exception:  # noqa: BLE001
                        extracted = None
                    if extracted:
                        tables.append(extracted)
                if found:
                    text = _text_outside_tables(p, found)
                else:
                    text = p.extract_text() or ""
                parts = [text.strip()] if text.strip() else []
                for t in tables:
                    rows = []
                    for row in t:
                        cells = [(c or "").replace("\n", " ").strip() for c in row]
                        if any(cells):
                            rows.append(" | ".join(cells))
                    if rows:
                        parts.append("\n".join(rows))
                    raw_cells = [[(c or "").replace("\n", " ").strip() for c in row] for row in t]
                    facts = _scheme_summary_facts(raw_cells)
                    if facts:
                        parts.append(facts)
                if parts:
                    pages.append(
                        RawPage(page=i, label="page", text="\n\n".join(parts), is_table=bool(tables))
                    )
        return pages
    except ImportError:
        logger.warning("pdfplumber not available; falling back to pypdf for %s", path.name)

    from pypdf import PdfReader

    reader = PdfReader(str(path))
    for i, p in enumerate(reader.pages, start=1):
        try:
            text = p.extract_text() or ""
        except Exception:
            text = ""
        if text.strip():
            pages.append(RawPage(page=i, label="page", text=text))
    return pages


def _sheet_to_rows(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if v is None else str(v).strip() for v in row]
        if any(cells):
            rows.append(cells)
    return rows


def extract_xlsx(path: Path) -> list[RawPage]:
    import io

    import openpyxl

    # Handle .xls files that are actually xlsx content (HDFC TER exports)
    data = None
    if path.suffix.lower() == ".xls":
        raw = path.read_bytes()[:4]
        if raw[:2] == b"PK":
            data = path.read_bytes()
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        else:
            raise ValueError(
                f"Legacy .xls format not supported for {path.name}: "
                "please convert to .xlsx"
            )
    else:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Detect TER docs: header contains Total/Base Expense Ratio
    is_ter = False
    try:
        first_ws = wb.worksheets[0]
        for row in first_ws.iter_rows(max_row=5, values_only=True):
            txt = " ".join(str(c) for c in row if c is not None).lower()
            if "total expense ratio" in txt or "base expense ratio" in txt:
                is_ter = True
                break
    except Exception:
        pass

    if is_ter:
        return _extract_ter_collapsed(wb)

    out: list[RawPage] = []
    for idx, ws in enumerate(wb.worksheets, start=1):
        rows = _sheet_to_rows(ws)
        if not rows:
            continue
        lines: list[str] = []
        header: list[str] | None = None
        for r in rows:
            # treat the first non-empty row as header for context
            if header is None:
                header = r
                lines.append("HEADER: " + " | ".join(header))
                continue
            cells = []
            for h, v in zip(header, r + [""] * max(0, len(header) - len(r))):
                if v:
                    cells.append(f"{h}: {v}" if h else v)
            lines.append(" || ".join(cells) if cells else "")
        text = "\n".join(line for line in lines if line)
        if text.strip():
            out.append(RawPage(page=idx, label=ws.title or f"sheet{idx}", text=text, is_table=True))
    wb.close()
    return out


def _extract_ter_collapsed(wb) -> list[RawPage]:
    """Collapse TER daily rows → latest row per scheme (one compact fact table).

    HDFC's TER sheet has duplicate column groups:
      cols 3-7  = Regular Plan (BER, Brokerage, Transaction, Levies, Total TER)
      cols 8-12 = Direct Plan  (BER, Brokerage, Transaction, Levies, Total TER)
    The only label distinguishing them is the merged title row that contains
    "Regular Plan" / "Direct Plan" above those columns. We rewrite the
    header to make each group explicit so that queries with or without
    "Direct Plan" can be answered from the same chunk.

    Output is one natural-language fact per scheme (one scheme = one chunk),
    so the Direct Plan value never gets cut away from its scheme name and
    the embedding is close to natural questions like
    "What is the expense ratio of HDFC FlexiCap Fund?".
    """
    from collections import defaultdict
    from datetime import datetime

    ws = wb.worksheets[0]
    rows = _sheet_to_rows(ws)
    if not rows:
        wb.close()
        return []

    # Find the real column-header row (contains "Scheme Name" + "Base Expense Ratio")
    header = None
    header_row_idx = 0
    for i, r in enumerate(rows):
        low = [h.lower() for h in r]
        if any("scheme name" in h for h in low) and any("base expense" in h for h in low):
            header = r
            header_row_idx = i
            break
    if header is None:
        wb.close()
        return []

    # Detect which column ranges are Regular vs Direct from the title row
    # that contains "Regular Plan" / "Direct Plan" merged cells. Fall back to
    # positional: first half Regular, second half Direct.
    regular_cols: set[int] = set()
    direct_cols: set[int] = set()
    if header_row_idx > 0:
        title_row = rows[header_row_idx - 1] if header_row_idx > 0 else []
        title_row += [""] * (len(header) - len(title_row))
        for idx, cell in enumerate(title_row):
            low = cell.lower()
            if "regular" in low:
                regular_cols.add(idx)
            elif "direct" in low:
                direct_cols.add(idx)
        # If title row has merged cells (only first cell of range has text),
        # propagate: all cols between Regular label and Direct label are Regular.
        if regular_cols and direct_cols:
            r_start = min(regular_cols)
            d_start = min(direct_cols)
            regular_cols = set(range(r_start, d_start))
            direct_cols = set(range(d_start, len(header)))

    # If no title row labels found, fall back to positional split
    if not regular_cols and not direct_cols:
        mid = len(header) // 2
        # Scheme Name + NSDL + Date are shared (cols 0-2)
        regular_cols = set(range(3, mid)) if mid > 3 else set()
        direct_cols = set(range(mid, len(header)))

    # Rewrite header to make plan explicit
    labeled_header = []
    for idx, h in enumerate(header):
        if idx in direct_cols:
            labeled_header.append(f"Direct Plan - {h}")
        elif idx in regular_cols:
            labeled_header.append(f"Regular Plan - {h}")
        else:
            labeled_header.append(h)

    # Find column indices (case-insensitive) on original header
    header_l = [h.lower() for h in header]
    try:
        scheme_idx = next(i for i, h in enumerate(header_l) if "scheme name" in h)
        date_idx = next(i for i, h in enumerate(header_l) if "date" in h)
    except StopIteration:
        wb.close()
        return []

    # Group by scheme, keep latest date
    latest: dict[str, tuple[datetime, list[str]]] = {}
    for r in rows[header_row_idx + 1:]:
        if scheme_idx >= len(r) or date_idx >= len(r):
            continue
        scheme = r[scheme_idx].strip()
        date_str = r[date_idx].strip()
        if not scheme or not date_str:
            continue
        d = None
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
        if d is None:
            continue
        if scheme not in latest or d > latest[scheme][0]:
            latest[scheme] = (d, r)

    # Return one RawPage per scheme so each scheme's Direct/Regular TER
    # is isolated and retrievable by scheme-filtered search, with per-chunk
    # scheme detection accurate. Use natural language for better embedding
    # similarity to questions like "What is the expense ratio of HDFC FlexiCap Fund?"
    labeled_header_l = [h.lower() for h in labeled_header]
    try:
        direct_ter_idx = next(i for i, h in enumerate(labeled_header_l) if "direct plan - total ter" in h)
    except StopIteration:
        direct_ter_idx = None
    try:
        regular_ter_idx = next(i for i, h in enumerate(labeled_header_l) if "regular plan - total ter" in h)
    except StopIteration:
        regular_ter_idx = None

    out: list[RawPage] = []
    for scheme in sorted(latest.keys()):
        r = latest[scheme][1]
        date_str = r[date_idx].strip() if date_idx < len(r) else ""
        # Build natural language fact
        parts = [f"The Total Expense Ratio (TER) of {scheme} as of {date_str} is"]
        details = []
        if regular_ter_idx is not None and regular_ter_idx < len(r) and r[regular_ter_idx].strip():
            details.append(f"{r[regular_ter_idx].strip()}% for Regular Plan")
        if direct_ter_idx is not None and direct_ter_idx < len(r) and r[direct_ter_idx].strip():
            details.append(f"{r[direct_ter_idx].strip()}% for Direct Plan")
        if details:
            parts.append(" and ".join(details) + ".")
        else:
            # Fallback: include all labeled cells
            cells = []
            for h, v in zip(labeled_header, r + [""] * max(0, len(labeled_header) - len(r))):
                if v:
                    cells.append(f"{h}: {v}" if h else v)
            parts.append(" || ".join(cells))

        # Add full labeled row for completeness (helps citation with all columns)
        cells = []
        for h, v in zip(labeled_header, r + [""] * max(0, len(labeled_header) - len(r))):
            if v:
                cells.append(f"{h}: {v}" if h else v)
        full_row = " || ".join(cells)

        text = " ".join(parts) + "\n" + full_row
        # Also include header for context
        text = f"Total Expense Ratio (TER) Report - {scheme}\n" + text
        out.append(RawPage(page=1, label="TER Report", text=text, is_table=True))
    wb.close()
    return out if out else []


def extract_file(path: Path) -> list[RawPage]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf(path)
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return extract_xlsx(path)
    if suffix in (".txt", ".md"):
        text = path.read_text(encoding="utf-8", errors="replace")
        return [RawPage(page=1, label="document", text=text)]
    raise ValueError(f"Unsupported file type: {suffix}")
