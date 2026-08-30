"""Bootstrap data/manifest.csv from the raw 'HDFC MF PDFs' source tree.

- Copies every supported file into data/documents/
- Infers scheme / organization / document_type from folder + filename
- Pulls official source URLs from HDFC_Mutual_Fund_FAQ.xlsx ('Source URLs' sheet)
  with sensible official fallbacks.

Usage:
    python scripts/bootstrap_manifest.py [--source-dir "path/to/HDFC MF PDFs"]
"""
from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from ingestion.config import (  # noqa: E402
    DEFAULT_SOURCE_URLS,
    detect_document_type,
    detect_organization,
    detect_scheme,
)

SCHEME_FOLDERS = {
    "elss tax": "HDFC ELSS Tax Saver Fund",
    "flexi cap": "HDFC FlexiCap Fund",
    "large and mid cap": "HDFC Large and Mid Cap Fund",
    "nifty50 index": "HDFC Index Fund - Nifty 50 Plan",
    "small cap": "HDFC Small Cap Fund",
}

SUPPORTED_EXT = {".pdf", ".xlsx", ".xlsm", ".txt", ".md", ".xls"}

DATE_FULL_RE = re.compile(
    r"(January|February|March|April|May|June|July|August|September|October|November|December|"
    r"Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sept|Sep|Oct|Nov|Dec)\s+(\d{1,2}),?\s+(\d{4})",
    re.IGNORECASE,
)
DATE_MY_RE = re.compile(
    r"(January|February|March|April|May|June|July|August|September|October|November|December|"
    r"Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sept|Sep|Oct|Nov|Dec)\s+(\d{2,4})",
    re.IGNORECASE,
)

_MONTH_NAMES = [
    "january", "february", "march", "april", "may", "june", "july",
    "august", "september", "october", "november", "december",
]
MONTHS = {m: i + 1 for i, m in enumerate(_MONTH_NAMES)}
for _i in range(1, 13):
    full = _MONTH_NAMES[_i - 1]
    MONTHS[full[:3]] = _i
MONTHS["sept"] = 9


def infer_date_from_name(name: str) -> str:
    """Return YYYY-MM-DD from 'November 21, 2025', 'July 2026', or 'July 26'."""
    m = DATE_FULL_RE.search(name)
    if m:
        month = MONTHS.get(m.group(1).lower())
        day, year = int(m.group(2)), int(m.group(3))
        if month and 2015 <= year <= 2035 and 1 <= day <= 31:
            return f"{year:04d}-{month:02d}-{day:02d}"
    m = DATE_MY_RE.search(name)
    if m:
        month = MONTHS.get(m.group(1).lower())
        year = int(m.group(2))
        if year < 100:
            year += 2000
        if month and 2015 <= year <= 2035:
            return f"{year:04d}-{month:02d}-01"
    return ""


def load_workbook_source_urls(faq_xlsx: Path) -> dict[str, str]:
    """scheme/organization -> official URL, from the 'Source URLs' sheet."""
    urls: dict[str, str] = {}
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(faq_xlsx), read_only=True, data_only=True)
        if "Source URLs" in wb.sheetnames:
            ws = wb["Source URLs"]
            rows = list(ws.iter_rows(values_only=True))
            header = [str(c or "").strip().lower() for c in rows[0]]
            org_i = header.index("organization") if "organization" in header else 0
            scope_i = header.index("scope") if "scope" in header else 1
            url_i = header.index("url") if "url" in header else 4
            for r in rows[1:]:
                if r[url_i]:
                    scope = str(r[scope_i] or "").strip()
                    urls[scope] = str(r[url_i]).strip()
                    urls.setdefault(str(r[org_i] or "").strip(), str(r[url_i]).strip())
        wb.close()
    except Exception as exc:  # noqa: BLE001
        print(f"  ! Could not read source URLs from workbook: {exc}")
    return urls


def infer_date_from_name_old_removed(name: str) -> str:  # kept for reference; superseded by infer_date_from_name above
    """Return YYYY-MM-DD (day=01) from patterns like 'July 2026' or 'July 26'."""
    m = DATE_MY_RE.search(name)
    if not m:
        return ""
    month = MONTHS.get(m.group(1).lower())
    year = int(m.group(2))
    if year < 100:
        year += 2000
    if not month or not (2015 <= year <= 2035):
        return ""
    return f"{year:04d}-{month:02d}-01"


def title_for(name: str, doc_type: str) -> str:
    stem = Path(name).stem.replace("_", " ").replace("-", " ").strip()
    return f"{stem} ({doc_type})"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source-dir",
        default=str(ROOT / "HDFC MF PDFs"),
        help="Directory containing the raw official documents",
    )
    args = parser.parse_args()

    src_dir = Path(args.source_dir)
    docs_dir = ROOT / "data" / "documents"
    docs_dir.mkdir(parents=True, exist_ok=True)
    (docs_dir / ".gitkeep").touch(exist_ok=True)

    faq_xlsx = src_dir / "HDFC_Mutual_Fund_FAQ.xlsx"
    wb_urls = load_workbook_source_urls(faq_xlsx) if faq_xlsx.exists() else {}

    rows = []
    seen_names: dict[str, int] = {}
    # The FAQ workbook is an evaluation asset (questions without answers) —
    # ingesting it would let the model "answer" from bare questions.
    EXCLUDE_FILES = {"hdfc_mutual_fund_faq.xlsx"}
    for path in sorted(src_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXT:
            continue
        if path.name.lower().replace(" ", "_") in EXCLUDE_FILES or path.name.lower() in EXCLUDE_FILES:
            print(f"  skipping eval asset: {path.name}")
            continue

        rel_parent = path.parent.name.lower()
        folder_scheme = next((v for k, v in SCHEME_FOLDERS.items() if k in rel_parent), "")
        name_text = path.stem
        scheme = folder_scheme or detect_scheme(name_text)
        organization = (
            "Groww" if "groww" in path.name.lower()
            else ("SEBI" if "sebi" in path.name.lower()
                  else ("AMFI" if "amfi" in path.name.lower()
                        else ("HDFC Mutual Fund" if scheme else detect_organization(path.name))))
        )
        doc_type = detect_document_type(path.name)

        # unique flat file name
        flat = re.sub(r"[^A-Za-z0-9._ ()\-]", "", path.name).strip()
        base = flat or path.stem
        n = seen_names.get(base, 0)
        seen_names[base] = n + 1
        if n:
            flat = f"{Path(base).stem}_{n}{Path(base).suffix}"

        dest = docs_dir / flat
        shutil.copy2(path, dest)

        # source URL resolution priority: workbook(scheme) > default map > blank
        # NOTE: workbook org-level URLs are intentionally NOT used for
        # unschemed docs — they point at one specific fund.
        source_url = ""
        if scheme and scheme in wb_urls:
            source_url = wb_urls[scheme]
        if not source_url and not scheme:
            source_url = DEFAULT_SOURCE_URLS.get(organization, "")
        if not source_url:
            source_url = DEFAULT_SOURCE_URLS.get(scheme, "")

        rows.append(
            {
                "file_name": flat,
                "title": title_for(path.stem, doc_type),
                "scheme": scheme,
                "organization": organization,
                "document_type": doc_type,
                "document_date": infer_date_from_name(path.name),
                "effective_date": "",
                "source_url": source_url,
                "source_id": "",
            }
        )

    # Preserve manually-added docs that live in data/documents but not in source tree
    # (e.g., TER report and capital-gains blog PDF added directly to data/documents)
    out_path = ROOT / "data" / "manifest.csv"
    if out_path.exists():
        try:
            existing_rows = list(csv.DictReader(open(out_path, encoding="utf-8-sig")))
            existing_names = {r["file_name"] for r in rows}
            for er in existing_rows:
                fn = er.get("file_name", "")
                if fn and fn not in existing_names and (docs_dir / fn).exists():
                    # Fix known source URLs for preserved docs
                    if "capital gain" in fn.lower():
                        er["source_url"] = "https://www.hdfcfund.com/learn/blog/how-get-capital-gain-statement-mutual-fund-schemes-india"
                        er["organization"] = "HDFC Mutual Fund"
                    if "ter" in fn.lower() and not er.get("source_url"):
                        er["source_url"] = "https://www.hdfcfund.com/statutory-disclosure/total-expense-ratio-of-mutual-fund-schemes/reports"
                    rows.append(er)
                    print(f"  preserving existing doc: {fn}")
        except Exception as exc:  # noqa: BLE001
            print(f"  ! could not preserve existing manifest rows: {exc}")

    # Keep only the newest TER report — daily snapshots would conflict
    ter_rows = [r for r in rows if r["document_type"] == "TER_DATA"]
    if len(ter_rows) > 1:
        def _ter_key(r):
            # HDFCMF_SCHEMES_TER_23-08-2026.xls → 2026-08-23
            import datetime as _dt

            m = re.search(r"(\d{2})-(\d{2})-(\d{4})", r["file_name"])
            if m:
                d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                try:
                    return _dt.date(y, mo, d)
                except ValueError:
                    pass
            return _dt.date.min

        latest = max(ter_rows, key=_ter_key)
        for r in ter_rows:
            if r is not latest:
                try:
                    (docs_dir / r["file_name"]).unlink()
                except FileNotFoundError:
                    pass
                print(f"  archiving older TER: {r['file_name']}")
        rows = [r for r in rows if r["document_type"] != "TER_DATA"] + [latest]

        # Fix TER-specific metadata that generic inference misses
        for r in rows:
            if r["document_type"] == "TER_DATA":
                r["organization"] = "HDFC Mutual Fund"
                r["source_url"] = "https://www.hdfcfund.com/statutory-disclosure/total-expense-ratio-of-mutual-fund-schemes/reports"
                r["title"] = f"HDFC Mutual Fund TER Report {r['document_date']} (TER_DATA)"
                # Try dd-mm-yyyy fallback for TER filenames like 23-08-2026
                if not r["document_date"]:
                    m2 = re.search(r"(\d{2})-(\d{2})-(\d{4})", r["file_name"])
                    if m2:
                        r["document_date"] = f"{m2.group(3)}-{m2.group(2)}-{m2.group(1)}"

    out_path = ROOT / "data" / "manifest.csv"
    fields = ["file_name", "title", "scheme", "organization", "document_type",
              "document_date", "effective_date", "source_url", "source_id"]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} files to {docs_dir}")
    print(f"Manifest written to {out_path}")
    for r in rows:
        flag = "" if r["source_url"] else "   [NO URL]"
        print(f"  {r['document_type']:<12} {r['scheme'] or r['organization']:<34} {r['file_name']}{flag}")


if __name__ == "__main__":
    main()
